import sys
from openpyxl import (load_workbook)

wb = load_workbook(sys.argv[1], data_only=True)
work_sheet = wb.get_active_sheet()

with open("cmd.txt", "w") as f:
    for col in range(work_sheet.min_column, work_sheet.max_column, 1):
        cell = work_sheet.cell(row=work_sheet.min_row, column=col)
        ne_name = ''
        if cell.value is not None:
            ne_name = cell.value
            f.write("starttrans;\n")
            f.write('''add /NE/%(ne_name)s;\n''' % {'ne_name': ne_name})
            for i in range(col, col + 2, 1):
                field_name = work_sheet.cell(row=work_sheet.min_row + 1, column=i).value
                for row_ in range(work_sheet.min_row + 2, 25, 1):
                    frequency = work_sheet.cell(row=row_, column=i).value
                    if frequency is not None and frequency != 0:
                        f.write('''add /NE/%(ne_name)s/%(field_name)s/%(frequency)s;\n''' % {'ne_name': ne_name,
                                                                                             'field_name': field_name,
                                                                                             'frequency': str(
                                                                                                 frequency)})
            f.write("commit;\n")
